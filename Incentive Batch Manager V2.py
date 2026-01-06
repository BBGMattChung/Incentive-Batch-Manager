"""
Incentive Batch Splitter & Mail Merge
Complete workflow: Split Master Incentive Log by SalesPersonID,
sync email lists, and distribute via Outlook.

Author: Josh Perry - TD Analyst, Breakthru Beverage Group
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from pathlib import Path
import os
import shutil
import tempfile
import random
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Outlook import - will fail gracefully if not on Windows
try:
    import win32com.client as win32
    OUTLOOK_AVAILABLE = True
except ImportError:
    OUTLOOK_AVAILABLE = False


class IncentiveBatchManager:
    def __init__(self, root):
        self.root = root
        self.root.title("Incentive Batch Manager")
        self.root.geometry("750x1100")
        self.root.resizable(False, False)
        
        # Configure styles
        self.style = ttk.Style()
        self.style.configure("Title.TLabel", font=("Segoe UI", 16, "bold"))
        self.style.configure("Section.TLabel", font=("Segoe UI", 11, "bold"))
        self.style.configure("Header.TLabel", font=("Segoe UI", 10, "bold"))
        self.style.configure("Info.TLabel", font=("Segoe UI", 9))
        self.style.configure("Success.TLabel", font=("Segoe UI", 9), foreground="green")
        self.style.configure("Warning.TLabel", font=("Segoe UI", 9), foreground="orange")
        self.style.configure("Error.TLabel", font=("Segoe UI", 9), foreground="red")
        
        # ===== File Paths =====
        self.source_path = Path(
            r"C:\Users\jmperry\Breakthru Beverage Group\BBG CA - TD Analysts - Documents"
            r"\General\Incentives In Progress\Completed Incentives\1. Master Log\Master Incentive Log.xlsx"
        )
        self.output_base = Path(
            r"C:\Users\jmperry\Breakthru Beverage Group\BBG CA - TD Analysts - Documents\Incentive Mail Merge"
        )
        self.email_list_path = Path(
            r"C:\Users\jmperry\Breakthru Beverage Group\BBG CA - TD Analysts - Documents"
            r"\Incentive Mail Merge\EmailList.xlsx"
        )
        self.hierarchy_path = Path(
            r"C:\Users\jmperry\Breakthru Beverage Group\BBG CA - TD Analysts - Documents"
            r"\Form Site\CA_Sales Detail.xlsx"
        )
        
        # Output columns for batch files
        self.output_columns = [
            'Incentive #', 'SalesPersonID', 'Position ID', 'Territory ID', 
            'Rep Name', 'Payout', 'Supplier', 'Desc', 'Sales Role', 'Channel', 
            'Payout Type', 'Start Date', 'End Date', 'Submitted By', 
            'Tracking Method', 'Paid On Batch', 'Batch Year'
        ]
        
        # Email settings
        self.sender_email = "CAFinance_Incentives@breakthrubev.com"
        self.cc_email = "CAFinance_Incentives@breakthrubev.com"
        self.recap_recipients = [
            "CAFinance_Incentives@breakthrubev.com",
            "CAAnalyticsAll@breakthrubev.com"
        ]
        
        # ===== Data Storage =====
        self.master_df = None
        self.email_df = None
        self.hierarchy_df = None  # Phase 3: CA_Sales_Detail
        self.batch_files = []
        self.email_mapping = {}
        self.missing_emails = []
        self.created_batch_folder = None
        self.selected_year = None
        self.selected_batch = None
        
        # Phase 3: Management reporting
        self.mgmt_reports = []
        self.mgmt_email_mapping = {}
        self.mgmt_missing_emails = []
        self.mgmt_output_folder = None
        
        # Build UI
        self.create_widgets()
        
        # Load data on startup
        self.root.after(100, self.initial_load)
    
    def create_widgets(self):
        """Create all UI widgets."""
        # Create scrollable canvas
        canvas = tk.Canvas(self.root)
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas, padding="15")
        
        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Enable mousewheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        main_frame = self.scrollable_frame
        
        # ===== Title =====
        ttk.Label(
            main_frame, 
            text="Incentive Batch Manager",
            style="Title.TLabel"
        ).pack(pady=(0, 5))
        
        ttk.Label(
            main_frame,
            text="Split & Email Distribution Tool",
            style="Info.TLabel"
        ).pack(pady=(0, 15))
        
        # ========================================
        # STAGE 1: BATCH SPLITTING
        # ========================================
        stage1_frame = ttk.LabelFrame(main_frame, text="STAGE 1: Create Batch Files", padding="10")
        stage1_frame.pack(fill=tk.X, pady=(0, 15))
        
        # --- Status ---
        status_frame = ttk.Frame(stage1_frame)
        status_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(status_frame, text="Data Status:", style="Header.TLabel").pack(anchor=tk.W)
        self.data_status = ttk.Label(
            status_frame,
            text="Loading...",
            style="Info.TLabel"
        )
        self.data_status.pack(anchor=tk.W)
        
        # --- Batch Selection ---
        selection_frame = ttk.LabelFrame(stage1_frame, text="Batch Selection", padding="10")
        selection_frame.pack(fill=tk.X, pady=(5, 10))
        
        # Batch Year
        year_frame = ttk.Frame(selection_frame)
        year_frame.pack(fill=tk.X, pady=(0, 8))
        
        ttk.Label(year_frame, text="Batch Year:", width=15, style="Header.TLabel").pack(side=tk.LEFT)
        self.year_var = tk.StringVar()
        self.year_combo = ttk.Combobox(
            year_frame, 
            textvariable=self.year_var, 
            state="readonly",
            width=25
        )
        self.year_combo.pack(side=tk.LEFT, padx=(10, 0))
        self.year_combo.bind("<<ComboboxSelected>>", self.on_year_selected)
        
        # Paid On Batch
        batch_frame = ttk.Frame(selection_frame)
        batch_frame.pack(fill=tk.X)
        
        ttk.Label(batch_frame, text="Paid On Batch:", width=15, style="Header.TLabel").pack(side=tk.LEFT)
        self.batch_var = tk.StringVar()
        self.batch_combo = ttk.Combobox(
            batch_frame, 
            textvariable=self.batch_var, 
            state="readonly",
            width=25
        )
        self.batch_combo.pack(side=tk.LEFT, padx=(10, 0))
        self.batch_combo.bind("<<ComboboxSelected>>", self.on_batch_selected)
        
        # --- Preview ---
        preview_frame = ttk.LabelFrame(stage1_frame, text="Preview", padding="10")
        preview_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.preview_label1 = ttk.Label(
            preview_frame,
            text="Select Batch Year and Paid On Batch, then click 'Run Analysis'",
            style="Info.TLabel"
        )
        self.preview_label1.pack(anchor=tk.W)
        
        self.preview_label2 = ttk.Label(
            preview_frame,
            text="",
            style="Info.TLabel"
        )
        self.preview_label2.pack(anchor=tk.W, pady=(2, 10))
        
        # Buttons
        btn_frame = ttk.Frame(preview_frame)
        btn_frame.pack(fill=tk.X)
        
        self.analyze_btn = ttk.Button(
            btn_frame,
            text="Run Analysis",
            command=self.run_analysis,
            width=18
        )
        self.analyze_btn.pack(side=tk.LEFT, padx=(0, 10))
        
        self.create_files_btn = ttk.Button(
            btn_frame,
            text="Create Files",
            command=self.create_batch_files,
            state=tk.DISABLED,
            width=18
        )
        self.create_files_btn.pack(side=tk.LEFT)
        
        # --- Progress ---
        progress_frame = ttk.LabelFrame(stage1_frame, text="Progress", padding="10")
        progress_frame.pack(fill=tk.X)
        
        self.stage1_progress_var = tk.DoubleVar()
        self.stage1_progress = ttk.Progressbar(
            progress_frame,
            variable=self.stage1_progress_var,
            maximum=100,
            mode='determinate'
        )
        self.stage1_progress.pack(fill=tk.X, pady=(0, 5))
        
        self.stage1_progress_label = ttk.Label(
            progress_frame,
            text="Ready",
            style="Info.TLabel"
        )
        self.stage1_progress_label.pack(anchor=tk.W)
        
        # ========================================
        # STAGE 2: EMAIL DISTRIBUTION
        # ========================================
        stage2_frame = ttk.LabelFrame(main_frame, text="STAGE 2: Email Distribution", padding="10")
        stage2_frame.pack(fill=tk.X, pady=(0, 15))
        
        # --- Email List Sync ---
        sync_frame = ttk.LabelFrame(stage2_frame, text="Email List Sync", padding="10")
        sync_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.email_status = ttk.Label(
            sync_frame,
            text="Email list not loaded",
            style="Info.TLabel"
        )
        self.email_status.pack(anchor=tk.W)
        
        # ZREPCHECKNEW
        zrep_frame = ttk.Frame(sync_frame)
        zrep_frame.pack(fill=tk.X, pady=(8, 5))
        
        ttk.Label(zrep_frame, text="ZREPCHECKNEW:", style="Header.TLabel").pack(side=tk.LEFT)
        
        self.zrep_label = ttk.Label(
            zrep_frame,
            text="(Optional) No file selected",
            style="Info.TLabel"
        )
        self.zrep_label.pack(side=tk.LEFT, padx=(10, 10))
        
        ttk.Button(
            zrep_frame,
            text="Browse...",
            command=self.browse_zrepcheck,
            width=10
        ).pack(side=tk.RIGHT)
        
        # Sync button row
        sync_btn_frame = ttk.Frame(sync_frame)
        sync_btn_frame.pack(fill=tk.X, pady=(5, 0))
        
        self.sync_btn = ttk.Button(
            sync_btn_frame,
            text="Sync Email List",
            command=self.sync_email_list,
            state=tk.DISABLED,
            width=18
        )
        self.sync_btn.pack(side=tk.LEFT)
        
        self.sync_status = ttk.Label(
            sync_btn_frame,
            text="",
            style="Info.TLabel"
        )
        self.sync_status.pack(side=tk.LEFT, padx=(10, 0))
        
        # --- Email Assessment ---
        assess_frame = ttk.LabelFrame(stage2_frame, text="Email Assessment", padding="10")
        assess_frame.pack(fill=tk.X, pady=(0, 10))
        
        self.assess_label1 = ttk.Label(
            assess_frame,
            text="Create batch files first, then run assessment",
            style="Info.TLabel"
        )
        self.assess_label1.pack(anchor=tk.W)
        
        self.assess_label2 = ttk.Label(
            assess_frame,
            text="",
            style="Info.TLabel",
            wraplength=680,
            justify=tk.LEFT
        )
        self.assess_label2.pack(anchor=tk.W, pady=(5, 10))
        
        self.assess_btn = ttk.Button(
            assess_frame,
            text="Run Email Assessment",
            command=self.run_email_assessment,
            state=tk.DISABLED,
            width=20
        )
        self.assess_btn.pack(anchor=tk.W)
        
        # --- Test Emails ---
        test_frame = ttk.LabelFrame(stage2_frame, text="Test Emails", padding="10")
        test_frame.pack(fill=tk.X, pady=(0, 10))
        
        test_addr_frame = ttk.Frame(test_frame)
        test_addr_frame.pack(fill=tk.X)
        
        ttk.Label(test_addr_frame, text="Test Email:", style="Header.TLabel").pack(side=tk.LEFT)
        
        self.test_email_var = tk.StringVar()
        self.test_email_entry = ttk.Entry(
            test_addr_frame,
            textvariable=self.test_email_var,
            width=35
        )
        self.test_email_entry.pack(side=tk.LEFT, padx=(10, 0))
        
        test_btn_frame = ttk.Frame(test_frame)
        test_btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.test_btn = ttk.Button(
            test_btn_frame,
            text="Send 3 Test Emails",
            command=self.send_test_emails,
            state=tk.DISABLED,
            width=18
        )
        self.test_btn.pack(side=tk.LEFT)
        
        self.test_status = ttk.Label(
            test_btn_frame,
            text="",
            style="Info.TLabel"
        )
        self.test_status.pack(side=tk.LEFT, padx=(10, 0))
        
        # --- Send Final ---
        send_frame = ttk.LabelFrame(stage2_frame, text="Send Final Emails", padding="10")
        send_frame.pack(fill=tk.X)
        
        self.stage2_progress_var = tk.DoubleVar()
        self.stage2_progress = ttk.Progressbar(
            send_frame,
            variable=self.stage2_progress_var,
            maximum=100,
            mode='determinate'
        )
        self.stage2_progress.pack(fill=tk.X, pady=(0, 5))
        
        self.stage2_progress_label = ttk.Label(
            send_frame,
            text="Ready",
            style="Info.TLabel"
        )
        self.stage2_progress_label.pack(anchor=tk.W)
        
        send_btn_frame = ttk.Frame(send_frame)
        send_btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.send_btn = ttk.Button(
            send_btn_frame,
            text="Send Final Emails",
            command=self.send_final_emails,
            state=tk.DISABLED,
            width=18
        )
        self.send_btn.pack(side=tk.LEFT)
        
        self.send_status = ttk.Label(
            send_btn_frame,
            text="",
            style="Info.TLabel"
        )
        self.send_status.pack(side=tk.LEFT, padx=(10, 0))
        
        # ========================================
        # STAGE 3: MANAGEMENT REPORTING
        # ========================================
        stage3_frame = ttk.LabelFrame(main_frame, text="STAGE 3: Management Reports", padding="10")
        stage3_frame.pack(fill=tk.X, pady=(0, 15))
        
        # --- Hierarchy Status ---
        hier_status_frame = ttk.Frame(stage3_frame)
        hier_status_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(hier_status_frame, text="Hierarchy Data (CA_Sales_Detail):", style="Header.TLabel").pack(anchor=tk.W)
        self.hierarchy_status = ttk.Label(
            hier_status_frame,
            text="Not loaded",
            style="Info.TLabel"
        )
        self.hierarchy_status.pack(anchor=tk.W)
        
        # --- Generate Reports ---
        gen_frame = ttk.LabelFrame(stage3_frame, text="Generate Management Reports", padding="10")
        gen_frame.pack(fill=tk.X, pady=(5, 10))
        
        ttk.Label(
            gen_frame,
            text="Creates rollup reports for FSMs, Area Managers, Directors, and VPs",
            style="Info.TLabel"
        ).pack(anchor=tk.W)
        
        gen_btn_frame = ttk.Frame(gen_frame)
        gen_btn_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.gen_mgmt_btn = ttk.Button(
            gen_btn_frame,
            text="Generate Mgmt Reports",
            command=self.generate_management_reports,
            state=tk.DISABLED,
            width=20
        )
        self.gen_mgmt_btn.pack(side=tk.LEFT)
        
        self.mgmt_gen_status = ttk.Label(
            gen_btn_frame,
            text="",
            style="Info.TLabel"
        )
        self.mgmt_gen_status.pack(side=tk.LEFT, padx=(10, 0))
        
        # Coverage label
        self.mgmt_coverage_label = ttk.Label(
            gen_frame,
            text="",
            style="Info.TLabel"
        )
        self.mgmt_coverage_label.pack(anchor=tk.W, pady=(5, 0))
        
        # --- Management Email Distribution ---
        mgmt_email_frame = ttk.LabelFrame(stage3_frame, text="Management Email Distribution", padding="10")
        mgmt_email_frame.pack(fill=tk.X, pady=(5, 10))
        
        mgmt_assess_frame = ttk.Frame(mgmt_email_frame)
        mgmt_assess_frame.pack(fill=tk.X, pady=(0, 5))
        
        self.mgmt_assess_btn = ttk.Button(
            mgmt_assess_frame,
            text="Assess Mgmt Emails",
            command=self.run_mgmt_email_assessment,
            state=tk.DISABLED,
            width=18
        )
        self.mgmt_assess_btn.pack(side=tk.LEFT)
        
        self.mgmt_assess_status = ttk.Label(
            mgmt_assess_frame,
            text="",
            style="Info.TLabel"
        )
        self.mgmt_assess_status.pack(side=tk.LEFT, padx=(10, 0))
        
        # Test email
        mgmt_test_frame = ttk.Frame(mgmt_email_frame)
        mgmt_test_frame.pack(fill=tk.X, pady=(5, 0))
        
        ttk.Label(mgmt_test_frame, text="Test Email:", style="Info.TLabel").pack(side=tk.LEFT)
        self.mgmt_test_email_var = tk.StringVar()
        self.mgmt_test_email_entry = ttk.Entry(
            mgmt_test_frame,
            textvariable=self.mgmt_test_email_var,
            width=30
        )
        self.mgmt_test_email_entry.pack(side=tk.LEFT, padx=(5, 10))
        
        self.mgmt_test_btn = ttk.Button(
            mgmt_test_frame,
            text="Send 3 Test Emails",
            command=self.send_mgmt_test_emails,
            state=tk.DISABLED,
            width=15
        )
        self.mgmt_test_btn.pack(side=tk.LEFT)
        
        # Final send
        mgmt_send_frame = ttk.Frame(mgmt_email_frame)
        mgmt_send_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.mgmt_send_btn = ttk.Button(
            mgmt_send_frame,
            text="Send Mgmt Emails",
            command=self.send_mgmt_final_emails,
            state=tk.DISABLED,
            width=18
        )
        self.mgmt_send_btn.pack(side=tk.LEFT)
        
        self.mgmt_send_status = ttk.Label(
            mgmt_send_frame,
            text="",
            style="Info.TLabel"
        )
        self.mgmt_send_status.pack(side=tk.LEFT, padx=(10, 0))
        
        # Progress
        self.stage3_progress = ttk.Progressbar(
            stage3_frame,
            mode='determinate',
            length=400
        )
        self.stage3_progress.pack(fill=tk.X, pady=(10, 0))
        
        self.stage3_progress_label = ttk.Label(
            stage3_frame,
            text="",
            style="Info.TLabel"
        )
        self.stage3_progress_label.pack(anchor=tk.W)
        
        # ===== Bottom Buttons =====
        bottom_frame = ttk.Frame(main_frame)
        bottom_frame.pack(fill=tk.X, pady=(10, 0))
        
        ttk.Button(
            bottom_frame,
            text="Reload All Data",
            command=self.initial_load,
            width=15
        ).pack(side=tk.LEFT, padx=(0, 10))
        
        ttk.Button(
            bottom_frame,
            text="Open Output Folder",
            command=self.open_output_folder,
            width=18
        ).pack(side=tk.LEFT)
        
        ttk.Button(
            bottom_frame,
            text="Exit",
            command=self.root.quit,
            width=12
        ).pack(side=tk.RIGHT)
    
    # ==========================================
    # DATA LOADING
    # ==========================================
    
    def clean_id(self, val):
        """Clean ID values - remove .0 decimals, strip whitespace."""
        if pd.isna(val):
            return None
        s = str(val).strip()
        # Remove .0 suffix from float conversion
        if s.endswith('.0'):
            s = s[:-2]
        # Try to normalize as integer
        try:
            s = str(int(float(s)))
        except (ValueError, TypeError):
            pass
        return s
    
    def initial_load(self):
        """Load all data on startup."""
        self.load_master_data()
        self.load_email_list()
        self.load_hierarchy_data()
        
        # Enable generate management reports button if data is loaded
        if self.master_df is not None and self.hierarchy_df is not None:
            self.gen_mgmt_btn.config(state=tk.NORMAL)
    
    def load_master_data(self):
        """Load the Master Incentive Log."""
        self.data_status.config(text="Loading Master Incentive Log...", style="Info.TLabel")
        self.root.update()
        
        try:
            if not self.source_path.exists():
                self.data_status.config(
                    text=f"Master file not found",
                    style="Error.TLabel"
                )
                return
            
            # Create temp copy to avoid file lock
            temp_dir = tempfile.gettempdir()
            temp_file = Path(temp_dir) / "temp_master_incentive_log.xlsx"
            
            try:
                shutil.copy2(self.source_path, temp_file)
                self.master_df = pd.read_excel(temp_file)
                temp_file.unlink()
            except PermissionError:
                self.data_status.config(
                    text="File locked - close Excel and click 'Reload All Data'",
                    style="Error.TLabel"
                )
                return
            
            # Get unique years
            year_values = sorted(self.master_df['Batch Year'].dropna().unique().tolist(), reverse=True)
            year_values = [str(int(y)) if isinstance(y, float) else str(y) for y in year_values]
            
            self.year_combo['values'] = year_values
            
            total_rows = len(self.master_df)
            self.data_status.config(
                text=f"✓ Loaded {total_rows:,} records | {len(year_values)} batch years",
                style="Success.TLabel"
            )
            
        except Exception as e:
            self.data_status.config(
                text=f"Error: {str(e)}",
                style="Error.TLabel"
            )
    
    def load_email_list(self):
        """Load the email list."""
        try:
            if not self.email_list_path.exists():
                self.email_status.config(
                    text="Email list not found",
                    style="Error.TLabel"
                )
                return
            
            self.email_df = pd.read_excel(self.email_list_path)
            
            # Clean ID columns (remove .0 decimals)
            id_columns = ['SU01 Acct #', 'HR Supervisor #', 'HR 2nd Line Manager #']
            for col in id_columns:
                if col in self.email_df.columns:
                    self.email_df[col] = self.email_df[col].apply(self.clean_id)
            
            record_count = len(self.email_df)
            unique_ids = self.email_df['SU01 Acct #'].nunique() if 'SU01 Acct #' in self.email_df.columns else 0
            
            self.email_status.config(
                text=f"✓ Loaded {record_count:,} records ({unique_ids:,} unique IDs)",
                style="Success.TLabel"
            )
            
        except Exception as e:
            self.email_status.config(
                text=f"Error: {str(e)}",
                style="Error.TLabel"
            )
    
    # ==========================================
    # STAGE 1: BATCH SPLITTING
    # ==========================================
    
    def on_year_selected(self, event=None):
        """Handle year selection."""
        selected_year = self.year_var.get()
        if not selected_year:
            return
        
        self.create_files_btn.config(state=tk.DISABLED)
        self.assess_btn.config(state=tk.DISABLED)
        self.test_btn.config(state=tk.DISABLED)
        self.send_btn.config(state=tk.DISABLED)
        
        # Filter batch values
        year_filter = self.master_df['Batch Year'].astype(str).str.contains(selected_year.split('.')[0])
        filtered_df = self.master_df[year_filter]
        
        batch_values = filtered_df['Paid On Batch'].dropna().unique().tolist()
        batch_values = sorted([str(b) for b in batch_values])
        
        self.batch_combo['values'] = batch_values
        self.batch_var.set('')
        
        self.preview_label1.config(text=f"{len(batch_values)} batch dates available for {selected_year}")
        self.preview_label2.config(text="")
    
    def on_batch_selected(self, event=None):
        """Handle batch selection."""
        self.create_files_btn.config(state=tk.DISABLED)
        self.preview_label1.config(text="Selection changed. Click 'Run Analysis' to preview.")
        self.preview_label2.config(text="")
    
    def filter_master_data(self, year, batch):
        """Filter master dataframe by year and batch."""
        year_match = self.master_df['Batch Year'].astype(str).str.startswith(year.split('.')[0])
        batch_match = self.master_df['Paid On Batch'].astype(str) == batch
        return self.master_df[year_match & batch_match]
    
    def run_analysis(self):
        """Run batch analysis/preview."""
        year = self.year_var.get()
        batch = self.batch_var.get()
        
        if not year or not batch:
            messagebox.showinfo("Selection Required", "Please select both Batch Year and Paid On Batch")
            return
        
        filtered_df = self.filter_master_data(year, batch)
        
        if filtered_df.empty:
            self.preview_label1.config(text="No records found for this selection", style="Warning.TLabel")
            self.preview_label2.config(text="")
            self.create_files_btn.config(state=tk.DISABLED)
            return
        
        unique_reps = filtered_df['SalesPersonID'].nunique()
        total_records = len(filtered_df)
        total_payout = filtered_df['Payout'].sum() if 'Payout' in filtered_df.columns else 0
        
        self.selected_year = year
        self.selected_batch = batch
        
        self.preview_label1.config(
            text=f"Records: {total_records:,} | Unique Reps: {unique_reps} | Total Payout: ${total_payout:,.2f}"
        )
        self.preview_label2.config(
            text=f"Ready to create {unique_reps} individual files. Click 'Create Files' to proceed."
        )
        
        self.create_files_btn.config(state=tk.NORMAL)
    
    def format_excel_file(self, filepath):
        """Format Excel file with styling."""
        wb = load_workbook(filepath)
        ws = wb.active
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        headers = {cell.value: cell.column for cell in ws[1]}
        
        if 'Payout' in headers:
            payout_col = headers['Payout']
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=payout_col).number_format = '"$"#,##0.00'
        
        for col_name in ['Start Date', 'End Date']:
            if col_name in headers:
                col_idx = headers[col_name]
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).number_format = 'M/D/YYYY'
        
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if max_length < cell_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 6, 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
    
    def format_summary_file(self, filepath):
        """Format summary Excel file."""
        wb = load_workbook(filepath)
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for sheet_name in ['Rep Summary', 'Batch Info']:
            ws = wb[sheet_name]
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            if sheet_name == 'Rep Summary':
                headers = {cell.value: cell.column for cell in ws[1]}
                if 'Payout' in headers:
                    payout_col = headers['Payout']
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=payout_col).number_format = '"$"#,##0.00'
            
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if max_length < cell_length:
                                max_length = cell_length
                    except:
                        pass
                adjusted_width = min(max_length + 6, 60)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
    
    def create_batch_files(self):
        """Create individual files per SalesPersonID."""
        year = self.selected_year
        batch = self.selected_batch
        
        if not year or not batch:
            return
        
        self.stage1_progress_var.set(0)
        self.stage1_progress_label.config(text="Starting...")
        self.root.update()
        
        try:
            filtered_df = self.filter_master_data(year, batch)
            
            if filtered_df.empty:
                messagebox.showwarning("No Data", "No records found")
                return
            
            # Create folder structure: year / batch / Individual Emails
            year_folder = self.output_base / str(year).split('.')[0]
            batch_folder_name = f"{batch}_{year}".replace("/", "-").replace("\\", "-")
            batch_folder = year_folder / batch_folder_name
            individual_folder = batch_folder / "Individual Emails"
            
            individual_folder.mkdir(parents=True, exist_ok=True)
            
            # Clear existing files in Individual Emails folder
            for existing_file in individual_folder.glob("*.xlsx"):
                existing_file.unlink()
            
            available_columns = [col for col in self.output_columns if col in filtered_df.columns]
            
            files_created = 0
            total_records_written = 0
            rep_summary = []
            
            grouped = list(filtered_df.groupby('SalesPersonID'))
            total_reps = len(grouped)
            
            for idx, (sales_person_id, group_df) in enumerate(grouped):
                progress_pct = ((idx + 1) / total_reps) * 100
                self.stage1_progress_var.set(progress_pct)
                self.stage1_progress_label.config(text=f"Creating file {idx + 1} of {total_reps}...")
                self.root.update()
                
                rep_name = group_df['Rep Name'].iloc[0] if 'Rep Name' in group_df.columns else "Unknown"
                rep_name_clean = str(rep_name).replace("/", "-").replace("\\", "-").replace(":", "-")
                
                batch_clean = str(batch).replace("/", "-").replace("\\", "-").replace(":", "-")
                year_clean = str(year).split('.')[0]
                
                filename = f"{sales_person_id}_{rep_name_clean}_{batch_clean}_{year_clean}.xlsx"
                filepath = individual_folder / filename
                
                output_df = group_df[available_columns].copy()
                output_df.to_excel(filepath, index=False, sheet_name='Incentive Data')
                
                self.format_excel_file(filepath)
                
                files_created += 1
                records = len(output_df)
                total_records_written += records
                payout = output_df['Payout'].sum() if 'Payout' in output_df.columns else 0
                
                rep_summary.append({
                    'SalesPersonID': sales_person_id,
                    'Rep Name': rep_name,
                    'Records': records,
                    'Payout': payout
                })
            
            # Keep summary_df for the message but don't create file
            summary_df = pd.DataFrame(rep_summary)
            
            self.stage1_progress_var.set(100)
            self.stage1_progress_label.config(text=f"✓ Complete! {files_created} files created.")
            
            # Store for Stage 2
            self.created_batch_folder = individual_folder
            self.batch_files = [f for f in individual_folder.glob("*.xlsx") if not f.name.startswith("_")]
            
            # Enable Stage 2
            self.assess_btn.config(state=tk.NORMAL)
            self.assess_label1.config(text=f"Batch folder ready: {individual_folder.name} ({len(self.batch_files)} files)")
            
            messagebox.showinfo(
                "Files Created",
                f"BATCH SPLIT COMPLETE\n\n"
                f"Files Created: {files_created}\n"
                f"Total Records: {total_records_written:,}\n"
                f"Total Payout: ${summary_df['Payout'].sum():,.2f}\n\n"
                f"Location:\n{individual_folder}"
            )
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to create files:\n{str(e)}")
            self.stage1_progress_label.config(text=f"Error: {str(e)}")
    
    # ==========================================
    # STAGE 2: EMAIL DISTRIBUTION
    # ==========================================
    
    def browse_zrepcheck(self):
        """Browse for ZREPCHECKNEW file."""
        filepath = filedialog.askopenfilename(
            title="Select ZREPCHECKNEW File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
            initialdir=str(self.output_base)
        )
        
        if filepath:
            self.zrepcheck_path = Path(filepath)
            self.zrep_label.config(text=self.zrepcheck_path.name)
            self.sync_btn.config(state=tk.NORMAL)
    
    def sync_email_list(self):
        """Sync ZREPCHECKNEW into EmailList."""
        if not hasattr(self, 'zrepcheck_path') or not self.zrepcheck_path.exists():
            messagebox.showerror("Error", "Please select a valid ZREPCHECKNEW file")
            return
        
        if self.email_df is None:
            messagebox.showerror("Error", "Email list not loaded")
            return
        
        try:
            self.sync_status.config(text="Syncing...", style="Info.TLabel")
            self.root.update()
            
            zrep_df = pd.read_excel(self.zrepcheck_path)
            
            # Clean ID columns in new data (remove .0 decimals)
            id_columns = ['SU01 Acct #', 'HR Supervisor #', 'HR 2nd Line Manager #']
            for col in id_columns:
                if col in zrep_df.columns:
                    zrep_df[col] = zrep_df[col].apply(self.clean_id)
            
            # Also ensure existing email_df has clean IDs
            for col in id_columns:
                if col in self.email_df.columns:
                    self.email_df[col] = self.email_df[col].apply(self.clean_id)
            
            initial_count = len(self.email_df)
            
            combined_df = pd.concat([self.email_df, zrep_df], ignore_index=True)
            combined_df = combined_df.drop_duplicates(subset=['SU01 Acct #'], keep='first')
            
            new_records = len(combined_df) - initial_count
            
            combined_df.to_excel(self.email_list_path, index=False)
            self.email_df = combined_df
            
            self.sync_status.config(
                text=f"✓ Added {new_records} new records",
                style="Success.TLabel"
            )
            
            self.email_status.config(
                text=f"✓ Loaded {len(combined_df):,} records ({combined_df['SU01 Acct #'].nunique():,} unique IDs)",
                style="Success.TLabel"
            )
            
            messagebox.showinfo("Sync Complete", f"Added {new_records} new records.\nTotal: {len(combined_df):,}")
            
        except Exception as e:
            self.sync_status.config(text=f"Error: {str(e)}", style="Error.TLabel")
    
    def run_email_assessment(self):
        """Run email assessment."""
        if not self.batch_files:
            messagebox.showinfo("No Files", "Create batch files first")
            return
        
        if self.email_df is None:
            messagebox.showinfo("No Email List", "Load email list first")
            return
        
        try:
            self.email_mapping = {}
            self.missing_emails = []
            
            total_files = len(self.batch_files)
            matched = 0
            missing = 0
            
            # Create lookup
            email_lookup = {}
            for _, row in self.email_df.iterrows():
                acct_id = str(row.get('SU01 Acct #', '')).strip()
                email = str(row.get('SU01 Email', '')).strip()
                name = str(row.get('SU01 Name', '')).strip()
                
                if acct_id and acct_id != 'nan':
                    email_lookup[acct_id] = {
                        'email': email if email and email != 'nan' else None,
                        'name': name if name and name != 'nan' else 'Unknown'
                    }
            
            for file_path in self.batch_files:
                filename = file_path.stem
                parts = filename.split('_')
                
                if parts:
                    sales_person_id = str(parts[0]).strip()
                    
                    if sales_person_id in email_lookup:
                        email_info = email_lookup[sales_person_id]
                        
                        if email_info['email'] and '@' in email_info['email']:
                            self.email_mapping[sales_person_id] = {
                                'file': file_path,
                                'email': email_info['email'],
                                'name': email_info['name']
                            }
                            matched += 1
                        else:
                            self.missing_emails.append({
                                'SalesPersonID': sales_person_id,
                                'File': file_path.name,
                                'Reason': 'Invalid/missing email'
                            })
                            missing += 1
                    else:
                        self.missing_emails.append({
                            'SalesPersonID': sales_person_id,
                            'File': file_path.name,
                            'Reason': 'ID not in email list'
                        })
                        missing += 1
            
            if missing == 0:
                self.assess_label1.config(
                    text=f"✓ All {total_files} files have valid email addresses!",
                    style="Success.TLabel"
                )
                self.assess_label2.config(text="Ready to send emails.")
            else:
                self.assess_label1.config(
                    text=f"⚠ {matched} of {total_files} files have valid emails. {missing} missing.",
                    style="Warning.TLabel"
                )
                
                missing_text = "Missing: "
                for item in self.missing_emails[:5]:
                    missing_text += f"{item['SalesPersonID']} ({item['Reason']}), "
                if len(self.missing_emails) > 5:
                    missing_text += f"... +{len(self.missing_emails) - 5} more"
                
                self.assess_label2.config(text=missing_text.rstrip(", "))
            
            if matched > 0:
                self.test_btn.config(state=tk.NORMAL)
                self.send_btn.config(state=tk.NORMAL)
            
        except Exception as e:
            self.assess_label1.config(text=f"Error: {str(e)}", style="Error.TLabel")
    
    def get_batch_display_info(self):
        """Parse batch info for email content."""
        if not self.selected_batch:
            return "Unknown", "Unknown"
        
        batch_date = self.selected_batch
        batch_year = self.selected_year or ""
        
        try:
            date_parts = batch_date.replace("-", "/").split("/")
            if len(date_parts) >= 1:
                month_num = int(date_parts[0])
                month_names = ['January', 'February', 'March', 'April', 'May', 'June',
                               'July', 'August', 'September', 'October', 'November', 'December']
                batch_month = month_names[month_num - 1] if 1 <= month_num <= 12 else batch_date
            else:
                batch_month = batch_date
        except:
            batch_month = batch_date
        
        return batch_month, batch_year
    
    def create_email(self, outlook, recipient, sales_person_id, sales_person_name, 
                     file_path, batch_month, batch_year, is_test=False, test_email=None):
        """Create an Outlook email."""
        mail = outlook.CreateItem(0)
        
        mail.SentOnBehalfOfName = self.sender_email
        
        if is_test:
            mail.To = test_email
        else:
            mail.To = recipient
            mail.CC = self.cc_email
        
        batch_display = f"{batch_month} {batch_year}".strip()
        mail.Subject = f"Incentive Detail Report and Payout Information - {batch_display} - {sales_person_id}"
        
        if is_test:
            mail.Subject = f"[TEST] {mail.Subject}"
        
        body = f"""{sales_person_name} ({sales_person_id}) -

Attached is your incentive detail report for the {batch_month} {batch_year} payment. This report details the specific incentives you earned during this pay period.

If you have questions about the calculation of your final payout or any other aspect of your incentive details, please contact the CA Analytics Team - CAAnalyticsAll@breakthrubev.com

-Analytics Team"""
        
        mail.Body = body
        mail.Attachments.Add(str(file_path))
        
        return mail
    
    def send_test_emails(self):
        """Send 3 random test emails."""
        if not OUTLOOK_AVAILABLE:
            messagebox.showerror("Outlook Not Available", "pywin32 is required for Outlook integration")
            return
        
        test_email = self.test_email_var.get().strip()
        
        if not test_email or '@' not in test_email:
            messagebox.showerror("Invalid Email", "Please enter a valid test email address")
            return
        
        if len(self.email_mapping) < 3:
            messagebox.showerror("Not Enough Files", "Need at least 3 files with valid emails")
            return
        
        try:
            self.test_status.config(text="Sending...", style="Info.TLabel")
            self.root.update()
            
            random_ids = random.sample(list(self.email_mapping.keys()), 3)
            batch_month, batch_year = self.get_batch_display_info()
            
            outlook = win32.Dispatch('Outlook.Application')
            
            for sales_person_id in random_ids:
                info = self.email_mapping[sales_person_id]
                
                mail = self.create_email(
                    outlook=outlook,
                    recipient=info['email'],
                    sales_person_id=sales_person_id,
                    sales_person_name=info['name'],
                    file_path=info['file'],
                    batch_month=batch_month,
                    batch_year=batch_year,
                    is_test=True,
                    test_email=test_email
                )
                
                mail.Send()
            
            self.test_status.config(
                text=f"✓ 3 test emails sent to {test_email}",
                style="Success.TLabel"
            )
            
            messagebox.showinfo(
                "Test Complete",
                f"3 test emails sent to:\n{test_email}\n\n"
                f"SalesPersonIDs:\n• {random_ids[0]}\n• {random_ids[1]}\n• {random_ids[2]}"
            )
            
        except Exception as e:
            self.test_status.config(text=f"Error: {str(e)}", style="Error.TLabel")
            messagebox.showerror("Test Error", f"Failed:\n{str(e)}")
    
    def send_final_emails(self):
        """Send all final emails."""
        if not OUTLOOK_AVAILABLE:
            messagebox.showerror("Outlook Not Available", "pywin32 is required")
            return
        
        if not self.email_mapping:
            messagebox.showerror("No Emails", "Run assessment first")
            return
        
        total_emails = len(self.email_mapping)
        
        if not messagebox.askyesno(
            "Confirm Send",
            f"Send {total_emails} emails?\n\n"
            f"Batch: {self.created_batch_folder.name if self.created_batch_folder else 'Unknown'}\n"
            f"CC: {self.cc_email}"
        ):
            return
        
        try:
            batch_month, batch_year = self.get_batch_display_info()
            outlook = win32.Dispatch('Outlook.Application')
            
            sent_emails = []
            failed_emails = []
            
            for idx, (sales_person_id, info) in enumerate(self.email_mapping.items()):
                progress_pct = ((idx + 1) / total_emails) * 100
                self.stage2_progress_var.set(progress_pct)
                self.stage2_progress_label.config(text=f"Sending {idx + 1} of {total_emails}...")
                self.root.update()
                
                try:
                    mail = self.create_email(
                        outlook=outlook,
                        recipient=info['email'],
                        sales_person_id=sales_person_id,
                        sales_person_name=info['name'],
                        file_path=info['file'],
                        batch_month=batch_month,
                        batch_year=batch_year,
                        is_test=False
                    )
                    
                    mail.Send()
                    
                    sent_emails.append({
                        'SalesPersonID': sales_person_id,
                        'Name': info['name'],
                        'Email': info['email'],
                        'File': info['file'].name,
                        'Status': '✓ Sent'
                    })
                    
                except Exception as e:
                    failed_emails.append({
                        'SalesPersonID': sales_person_id,
                        'Name': info['name'],
                        'Email': info['email'],
                        'File': info['file'].name,
                        'Status': f'✗ Failed: {str(e)}'
                    })
            
            self.stage2_progress_var.set(100)
            self.stage2_progress_label.config(text=f"✓ Complete! {len(sent_emails)} sent.")
            
            self.send_recap_email(outlook, sent_emails, failed_emails, batch_month, batch_year)
            
            self.send_status.config(
                text=f"✓ {len(sent_emails)} sent, {len(failed_emails)} failed",
                style="Success.TLabel" if not failed_emails else "Warning.TLabel"
            )
            
            messagebox.showinfo(
                "Send Complete",
                f"Email distribution complete!\n\n"
                f"Sent: {len(sent_emails)}\n"
                f"Failed: {len(failed_emails)}\n\n"
                f"Recap sent to:\n• {self.recap_recipients[0]}\n• {self.recap_recipients[1]}"
            )
            
        except Exception as e:
            self.stage2_progress_label.config(text=f"Error: {str(e)}")
            messagebox.showerror("Send Error", f"Failed:\n{str(e)}")
    
    def send_recap_email(self, outlook, sent_emails, failed_emails, batch_month, batch_year):
        """Send recap email."""
        try:
            mail = outlook.CreateItem(0)
            
            mail.SentOnBehalfOfName = self.sender_email
            mail.To = "; ".join(self.recap_recipients)
            
            batch_display = f"{batch_month} {batch_year}".strip()
            mail.Subject = f"Incentive Email Distribution Recap - {batch_display}"
            
            total_sent = len(sent_emails)
            total_failed = len(failed_emails)
            success_rate = (total_sent / (total_sent + total_failed) * 100) if (total_sent + total_failed) > 0 else 0
            
            body = f"""INCENTIVE EMAIL DISTRIBUTION RECAP
{'='*50}

Batch: {self.created_batch_folder.name if self.created_batch_folder else 'Unknown'}
Date Sent: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

SUMMARY
{'-'*30}
Total Emails Sent: {total_sent}
Total Failed: {total_failed}
Success Rate: {success_rate:.1f}%


DETAILED SEND LOG
{'-'*30}
"""
            
            for item in sent_emails:
                body += f"✓ {item['SalesPersonID']} | {item['Name']} | {item['Email']}\n"
            
            if failed_emails:
                body += f"""

FAILED SENDS
{'-'*30}
"""
                for item in failed_emails:
                    body += f"✗ {item['SalesPersonID']} | {item['Name']} | {item['Status']}\n"
            
            body += f"""

{'='*50}
Automated recap from Incentive Batch Manager
"""
            
            mail.Body = body
            mail.Send()
            
        except Exception as e:
            messagebox.showwarning("Recap Warning", f"Emails sent but recap failed:\n{str(e)}")
    
    def open_output_folder(self):
        """Open the output folder."""
        if self.created_batch_folder and self.created_batch_folder.exists():
            os.startfile(self.created_batch_folder)
        else:
            os.startfile(self.output_base)
    
    # ==========================================
    # STAGE 3: MANAGEMENT REPORTING
    # ==========================================
    
    def load_hierarchy_data(self):
        """Load the CA_Sales_Detail hierarchy file."""
        try:
            if not self.hierarchy_path.exists():
                self.hierarchy_status.config(
                    text="Hierarchy file not found",
                    style="Error.TLabel"
                )
                return False
            
            self.hierarchy_df = pd.read_excel(self.hierarchy_path)
            
            # Clean all ID columns
            id_columns = [
                'Sales Person ID', 'Field Sales Manager ID', 
                'Area Manager ID', 'Division Manager ID', 'Vice President ID'
            ]
            for col in id_columns:
                if col in self.hierarchy_df.columns:
                    self.hierarchy_df[col] = self.hierarchy_df[col].apply(self.clean_id)
            
            # Filter to valid Division Manager Position IDs
            valid_dm_positions = ['40064102', '40064095', '40064096', '40064098', '40064471', '40064097', '40070873']
            self.hierarchy_df['Division Manager Position ID'] = self.hierarchy_df['Division Manager Position ID'].astype(str).str.strip()
            before_count = len(self.hierarchy_df)
            self.hierarchy_df = self.hierarchy_df[
                self.hierarchy_df['Division Manager Position ID'].isin(valid_dm_positions)
            ].copy()
            
            # Filter out unassigned FSMs
            self.hierarchy_df = self.hierarchy_df[
                self.hierarchy_df['Field Sales Manager ID'].notna()
            ].copy()
            
            self.hierarchy_status.config(
                text=f"✓ Loaded {len(self.hierarchy_df):,} records",
                style="Success.TLabel"
            )
            
            return True
            
        except Exception as e:
            self.hierarchy_status.config(
                text=f"Error: {str(e)}",
                style="Error.TLabel"
            )
            return False
    
    def generate_management_reports(self):
        """Generate management rollup reports."""
        if self.master_df is None or self.hierarchy_df is None:
            messagebox.showinfo("Data Required", "Please ensure all data is loaded")
            return
        
        year = self.year_var.get()
        batch = self.batch_var.get()
        
        if not year or not batch:
            messagebox.showinfo("Selection Required", "Please select a batch first in Stage 1")
            return
        
        try:
            self.mgmt_gen_status.config(text="Generating...", style="Info.TLabel")
            self.root.update()
            
            # Filter batch data
            batch_df = self.filter_master_data(year, batch).copy()
            batch_df['SalesPersonID'] = batch_df['SalesPersonID'].apply(self.clean_id)
            
            # Get incentive IDs
            incentive_ids = set(x for x in batch_df['SalesPersonID'].unique() if x is not None)
            hierarchy_ids = set(x for x in self.hierarchy_df['Sales Person ID'].dropna().unique() if x is not None)
            
            matched_ids = incentive_ids & hierarchy_ids
            unmatched_ids = incentive_ids - hierarchy_ids
            
            coverage_pct = len(matched_ids) / len(incentive_ids) * 100 if incentive_ids else 0
            self.mgmt_coverage_label.config(
                text=f"Coverage: {len(matched_ids)} of {len(incentive_ids)} matched ({coverage_pct:.1f}%), {len(unmatched_ids)} unmatched",
                style="Success.TLabel" if coverage_pct > 90 else "Warning.TLabel"
            )
            
            # Create output folder: year / batch / Management Reports
            year_clean = str(year).split('.')[0]
            batch_folder_name = f"{batch}_{year}".replace("/", "-").replace("\\", "-")
            output_folder = self.output_base / year_clean / batch_folder_name / "Management Reports"
            output_folder.mkdir(parents=True, exist_ok=True)
            
            # Clear old reports
            for f in output_folder.glob("*.xlsx"):
                f.unlink()
            
            self.mgmt_reports = []
            
            # Manager levels
            levels = [
                ('VP', 'Vice President ID', 'Vice President Name'),
                ('DIRECTOR', 'Division Manager ID', 'Division Manager Name'),
                ('AM', 'Area Manager ID', 'Area Manager Name'),
                ('FSM', 'Field Sales Manager ID', 'Field Sales Manager Name')
            ]
            
            total_managers = 0
            for level_code, id_col, name_col in levels:
                managers = self.hierarchy_df[[id_col, name_col]].drop_duplicates()
                for _, row in managers.iterrows():
                    mgr_id = row[id_col]
                    if pd.isna(mgr_id) or mgr_id in ['UNASSIGNED', None]:
                        continue
                    
                    # Check if manager has anyone with incentives
                    team = self.hierarchy_df[self.hierarchy_df[id_col] == mgr_id]
                    team_ids = set(team['Sales Person ID'].dropna().unique())
                    
                    if team_ids & incentive_ids:
                        total_managers += 1
            
            self.stage3_progress['maximum'] = total_managers
            self.stage3_progress['value'] = 0
            
            created = 0
            for level_code, id_col, name_col in levels:
                managers = self.hierarchy_df[[id_col, name_col]].drop_duplicates()
                
                for _, row in managers.iterrows():
                    mgr_id = row[id_col]
                    mgr_name = row[name_col]
                    
                    if pd.isna(mgr_id) or mgr_id in ['UNASSIGNED', None]:
                        continue
                    
                    # Check if manager has anyone with incentives
                    team = self.hierarchy_df[self.hierarchy_df[id_col] == mgr_id]
                    team_ids = set(team['Sales Person ID'].dropna().unique())
                    
                    if not (team_ids & incentive_ids):
                        continue
                    
                    filepath = self.create_manager_report(
                        level_code, mgr_id, mgr_name,
                        batch_df, year, batch, output_folder
                    )
                    
                    if filepath:
                        self.mgmt_reports.append({
                            'level': level_code,
                            'id': mgr_id,
                            'name': mgr_name,
                            'filepath': filepath
                        })
                        created += 1
                    
                    self.stage3_progress['value'] += 1
                    self.stage3_progress_label.config(text=f"Creating report {created}...")
                    self.root.update()
            
            self.mgmt_gen_status.config(
                text=f"✓ Created {created} reports",
                style="Success.TLabel"
            )
            self.stage3_progress_label.config(text=f"Complete: {created} management reports")
            
            # Enable assessment button
            if self.mgmt_reports:
                self.mgmt_assess_btn.config(state=tk.NORMAL)
            
            # Store output folder
            self.mgmt_output_folder = output_folder
            
            # Create Master Management Summary Report
            self.create_master_summary_report(batch_df, year, batch, output_folder)
            
        except Exception as e:
            self.mgmt_gen_status.config(text=f"Error: {str(e)}", style="Error.TLabel")
            messagebox.showerror("Error", f"Failed to generate reports:\n{str(e)}")
    
    def create_manager_report(self, manager_level, manager_id, manager_name, batch_df, batch_year, paid_on_batch, output_folder):
        """Create a multi-tab report for a manager."""
        level_config = {
            'VP': {'id_col': 'Vice President ID', 'include_levels': ['DIRECTOR', 'AM', 'FSM', 'REP']},
            'DIRECTOR': {'id_col': 'Division Manager ID', 'include_levels': ['AM', 'FSM', 'REP']},
            'AM': {'id_col': 'Area Manager ID', 'include_levels': ['FSM', 'REP']},
            'FSM': {'id_col': 'Field Sales Manager ID', 'include_levels': ['REP']}
        }
        
        config = level_config[manager_level]
        
        team_hierarchy = self.hierarchy_df[
            self.hierarchy_df[config['id_col']] == manager_id
        ].copy()
        
        if team_hierarchy.empty:
            return None
        
        level_id_cols = {
            'DIRECTOR': 'Division Manager ID', 'AM': 'Area Manager ID',
            'FSM': 'Field Sales Manager ID', 'REP': 'Sales Person ID'
        }
        level_labels = {'DIRECTOR': 'Director', 'AM': 'Area Mgr', 'FSM': 'FSM', 'REP': 'Rep'}
        
        # Hierarchy level order (for determining what columns to show)
        level_order = ['VP', 'DIRECTOR', 'AM', 'FSM', 'REP']
        manager_level_idx = level_order.index(manager_level)
        
        # Build lookups for each level type
        # For REPs: lookup by Sales Person ID
        rep_lookup = {}
        for _, row in team_hierarchy.iterrows():
            rep_id = row.get('Sales Person ID')
            if rep_id:
                rep_lookup[rep_id] = {
                    'Director': row.get('Division Manager Name', ''),
                    'Area Manager': row.get('Area Manager Name', ''),
                    'FSM': row.get('Field Sales Manager Name', '')
                }
        
        # For FSMs: lookup by Field Sales Manager ID
        fsm_lookup = {}
        for _, row in team_hierarchy.drop_duplicates(subset=['Field Sales Manager ID']).iterrows():
            fsm_id = row.get('Field Sales Manager ID')
            if fsm_id:
                fsm_lookup[fsm_id] = {
                    'Director': row.get('Division Manager Name', ''),
                    'Area Manager': row.get('Area Manager Name', '')
                }
        
        # For AMs: lookup by Area Manager ID
        am_lookup = {}
        for _, row in team_hierarchy.drop_duplicates(subset=['Area Manager ID']).iterrows():
            am_id = row.get('Area Manager ID')
            if am_id:
                am_lookup[am_id] = {
                    'Director': row.get('Division Manager Name', '')
                }
        
        # For Directors: no lookup needed (no hierarchy above within report)
        
        detail_columns = [
            'Incentive #', 'SalesPersonID', 'Position ID', 'Territory ID', 
            'Rep Name', 'Payout', 'Supplier', 'Desc', 'Sales Role', 
            'Channel', 'Payout Type', 'Start Date', 'End Date', 
            'Submitted By', 'Tracking Method', 'Paid On Batch', 'Batch Year'
        ]
        
        sheets_data = {}
        total_payout = 0
        total_people = 0
        
        for level in config['include_levels']:
            id_col = level_id_cols[level]
            label = level_labels[level]
            
            people_at_level = team_hierarchy[[id_col]].drop_duplicates()
            people_ids = [x for x in people_at_level[id_col].dropna().unique() if x is not None]
            
            if not people_ids:
                continue
            
            level_detail = batch_df[batch_df['SalesPersonID'].isin(people_ids)].copy()
            
            if level_detail.empty:
                continue
            
            # Determine which hierarchy columns to include for this tab
            # Include levels BETWEEN manager_level and current tab level
            current_level_idx = level_order.index(level)
            hier_cols_to_include = []
            
            for hier_level in ['DIRECTOR', 'AM', 'FSM']:
                hier_level_idx = level_order.index(hier_level)
                # Include if it's between manager level and current tab level
                if hier_level_idx > manager_level_idx and hier_level_idx < current_level_idx:
                    if hier_level == 'DIRECTOR':
                        hier_cols_to_include.append('Director')
                    elif hier_level == 'AM':
                        hier_cols_to_include.append('Area Manager')
                    elif hier_level == 'FSM':
                        hier_cols_to_include.append('FSM')
            
            # Select the right lookup based on the level we're reporting on
            if level == 'REP':
                lookup = rep_lookup
            elif level == 'FSM':
                lookup = fsm_lookup
            elif level == 'AM':
                lookup = am_lookup
            else:
                lookup = {}
            
            # Add hierarchy columns to detail
            for col_name in hier_cols_to_include:
                level_detail[col_name] = level_detail['SalesPersonID'].apply(
                    lambda x: lookup.get(x, {}).get(col_name, '')
                )
            
            # Summary - aggregate
            agg_dict = {'Rep Name': 'first', 'Payout': 'sum', 'Incentive #': 'count'}
            for col in hier_cols_to_include:
                agg_dict[col] = 'first'
            
            level_summary = level_detail.groupby('SalesPersonID').agg(agg_dict).reset_index()
            
            # Rename columns
            rename_cols = {'SalesPersonID': 'ID', 'Rep Name': 'Name', 'Payout': 'Total Payout', 'Incentive #': 'Incentive Count'}
            level_summary = level_summary.rename(columns=rename_cols)
            
            # Reorder: hierarchy cols FIRST, then ID, Name, Total Payout, Incentive Count
            col_order = hier_cols_to_include + ['ID', 'Name', 'Total Payout', 'Incentive Count']
            level_summary = level_summary[col_order]
            level_summary = level_summary.sort_values('Total Payout', ascending=False)
            
            # Detail - hierarchy columns first
            available_detail_cols = [col for col in detail_columns if col in level_detail.columns]
            level_detail_out = level_detail[hier_cols_to_include + available_detail_cols].copy()
            level_detail_out = level_detail_out.sort_values(['Rep Name', 'Payout'], ascending=[True, False])
            
            sheets_data[f'{label} Summary'] = level_summary
            sheets_data[f'{label} Detail'] = level_detail_out
            
            total_payout += level_summary['Total Payout'].sum()
            total_people += len(level_summary)
        
        if not sheets_data:
            return None
        
        manager_name_clean = str(manager_name).replace("/", "-").replace("\\", "-").replace(":", "-").replace("*", "-").replace("?", "-").strip()
        filename = f"{manager_level}_{manager_id}_{manager_name_clean}.xlsx"
        filepath = output_folder / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            summary_data = pd.DataFrame({
                'Metric': ['Manager Level', 'Manager ID', 'Manager Name', 'Batch Period',
                          'Total People with Incentives', 'Total Payout', 'Report Generated'],
                'Value': [manager_level, manager_id, manager_name, f"{paid_on_batch} {batch_year}",
                         total_people, f"${total_payout:,.2f}", datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            })
            summary_data.to_excel(writer, sheet_name='Summary', index=False)
            
            for sheet_name, df in sheets_data.items():
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
        
        # Apply formatting (auto-width, header color, currency)
        self.format_manager_report(filepath)
        
        return filepath
    
    def format_manager_report(self, filepath):
        """Format management report with auto-width, header color, and currency."""
        wb = load_workbook(filepath)
        
        # Match individual report header color: #4472C4
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        currency_format = '"$"#,##0.00'
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Header formatting
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Find payout columns and format as currency
            headers = {cell.value: cell.column for cell in ws[1]}
            
            payout_cols = [col for name, col in headers.items() 
                          if name and ('payout' in str(name).lower())]
            
            for col_idx in payout_cols:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).number_format = currency_format
            
            # Find date columns and format as short date
            date_cols = [col for name, col in headers.items() 
                        if name and ('start date' in str(name).lower() or 'end date' in str(name).lower())]
            
            for col_idx in date_cols:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col_idx).number_format = 'MM/DD/YYYY'
            
            # Auto-width columns
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if max_length < cell_length:
                                max_length = cell_length
                    except:
                        pass
                
                # Add padding, cap at 60
                adjusted_width = min(max_length + 6, 60)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(filepath)
    
    def create_master_summary_report(self, batch_df, batch_year, paid_on_batch, output_folder):
        """Create a master summary report with all levels and VP hierarchy for filtering."""
        try:
            # Build complete hierarchy lookup for all people
            hierarchy_lookup = {}
            for _, row in self.hierarchy_df.iterrows():
                for id_col in ['Sales Person ID', 'Field Sales Manager ID', 'Area Manager ID', 'Division Manager ID']:
                    person_id = row.get(id_col)
                    if person_id and person_id not in hierarchy_lookup:
                        hierarchy_lookup[person_id] = {
                            'VP': row.get('Vice President Name', ''),
                            'Director': row.get('Division Manager Name', ''),
                            'Area Manager': row.get('Area Manager Name', ''),
                            'FSM': row.get('Field Sales Manager Name', '')
                        }
            
            detail_columns = [
                'Incentive #', 'SalesPersonID', 'Position ID', 'Territory ID', 
                'Rep Name', 'Payout', 'Supplier', 'Desc', 'Sales Role', 
                'Channel', 'Payout Type', 'Start Date', 'End Date', 
                'Submitted By', 'Tracking Method', 'Paid On Batch', 'Batch Year'
            ]
            
            level_id_cols = {
                'DIRECTOR': 'Division Manager ID', 'AM': 'Area Manager ID',
                'FSM': 'Field Sales Manager ID', 'REP': 'Sales Person ID'
            }
            level_labels = {'DIRECTOR': 'Director', 'AM': 'Area Mgr', 'FSM': 'FSM', 'REP': 'Rep'}
            
            sheets_data = {}
            
            for level in ['DIRECTOR', 'AM', 'FSM', 'REP']:
                id_col = level_id_cols[level]
                label = level_labels[level]
                
                # Get all unique people at this level
                people_at_level = self.hierarchy_df[[id_col]].drop_duplicates()
                people_ids = [x for x in people_at_level[id_col].dropna().unique() if x is not None]
                
                if not people_ids:
                    continue
                
                # Get their incentive records
                level_detail = batch_df[batch_df['SalesPersonID'].isin(people_ids)].copy()
                
                if level_detail.empty:
                    continue
                
                # Add all hierarchy columns (including VP)
                hier_cols = ['VP', 'Director', 'Area Manager', 'FSM']
                for col_name in hier_cols:
                    level_detail[col_name] = level_detail['SalesPersonID'].apply(
                        lambda x: hierarchy_lookup.get(x, {}).get(col_name, '')
                    )
                
                # Determine which hierarchy columns to show (levels above current)
                if level == 'DIRECTOR':
                    show_hier = ['VP']
                elif level == 'AM':
                    show_hier = ['VP', 'Director']
                elif level == 'FSM':
                    show_hier = ['VP', 'Director', 'Area Manager']
                else:  # REP
                    show_hier = ['VP', 'Director', 'Area Manager', 'FSM']
                
                # Summary
                agg_dict = {'Rep Name': 'first', 'Payout': 'sum', 'Incentive #': 'count'}
                for col in show_hier:
                    agg_dict[col] = 'first'
                
                level_summary = level_detail.groupby('SalesPersonID').agg(agg_dict).reset_index()
                level_summary = level_summary.rename(columns={
                    'SalesPersonID': 'ID', 'Rep Name': 'Name', 
                    'Payout': 'Total Payout', 'Incentive #': 'Incentive Count'
                })
                
                col_order = show_hier + ['ID', 'Name', 'Total Payout', 'Incentive Count']
                level_summary = level_summary[col_order]
                level_summary = level_summary.sort_values('Total Payout', ascending=False)
                
                # Detail
                available_detail_cols = [col for col in detail_columns if col in level_detail.columns]
                level_detail_out = level_detail[show_hier + available_detail_cols].copy()
                level_detail_out = level_detail_out.sort_values(['Rep Name', 'Payout'], ascending=[True, False])
                
                sheets_data[f'{label} Summary'] = level_summary
                sheets_data[f'{label} Detail'] = level_detail_out
            
            if not sheets_data:
                return None
            
            # Calculate totals
            total_payout = batch_df['Payout'].sum()
            total_people = batch_df['SalesPersonID'].nunique()
            
            filepath = output_folder / "_MANAGEMENT_SUMMARY.xlsx"
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Summary sheet
                summary_data = pd.DataFrame({
                    'Metric': ['Report Type', 'Batch Period', 'Total People with Incentives', 
                              'Total Payout', 'Report Generated'],
                    'Value': ['Master Management Summary', f"{paid_on_batch} {batch_year}",
                             total_people, f"${total_payout:,.2f}", 
                             datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
                })
                summary_data.to_excel(writer, sheet_name='Summary', index=False)
                
                for sheet_name, df in sheets_data.items():
                    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
            
            self.format_manager_report(filepath)
            
        except Exception as e:
            print(f"Error creating master summary: {e}")
    
    def run_mgmt_email_assessment(self):
        """Assess email coverage for management reports."""
        if not self.mgmt_reports:
            messagebox.showinfo("No Reports", "Generate management reports first")
            return
        
        if self.email_df is None:
            messagebox.showinfo("No Email List", "Email list not loaded")
            return
        
        try:
            self.mgmt_email_mapping = {}
            self.mgmt_missing_emails = []
            
            for report in self.mgmt_reports:
                mgr_id = report['id']
                mgr_name = report['name']
                
                # Look up email
                match = self.email_df[self.email_df['SU01 Acct #'] == mgr_id]
                
                if not match.empty:
                    email = match.iloc[0].get('SU01 Email', '')
                    if email and '@' in str(email):
                        self.mgmt_email_mapping[mgr_id] = {
                            'email': email,
                            'name': mgr_name,
                            'filepath': report['filepath'],
                            'level': report['level']
                        }
                    else:
                        self.mgmt_missing_emails.append({
                            'id': mgr_id,
                            'name': mgr_name,
                            'reason': 'Invalid email'
                        })
                else:
                    self.mgmt_missing_emails.append({
                        'id': mgr_id,
                        'name': mgr_name,
                        'reason': 'ID not found in email list'
                    })
            
            matched = len(self.mgmt_email_mapping)
            total = len(self.mgmt_reports)
            
            if matched == total:
                self.mgmt_assess_status.config(
                    text=f"✓ All {total} managers have valid emails",
                    style="Success.TLabel"
                )
            else:
                self.mgmt_assess_status.config(
                    text=f"⚠ {matched} of {total} have emails. {len(self.mgmt_missing_emails)} missing",
                    style="Warning.TLabel"
                )
                
                # Show details of missing emails
                missing_details = "The following managers are missing valid emails:\n\n"
                for item in self.mgmt_missing_emails:
                    missing_details += f"• {item['id']} - {item['name']} ({item['reason']})\n"
                
                messagebox.showwarning("Missing Emails", missing_details)
            
            if self.mgmt_email_mapping:
                self.mgmt_test_btn.config(state=tk.NORMAL)
                self.mgmt_send_btn.config(state=tk.NORMAL)
            
        except Exception as e:
            self.mgmt_assess_status.config(text=f"Error: {str(e)}", style="Error.TLabel")
    
    def create_mgmt_email(self, outlook, recipient, manager_id, manager_name, manager_level, 
                          batch_month, batch_year, attachment_path, test_email=None):
        """Create a management report email."""
        mail = outlook.CreateItem(0)
        
        level_titles = {
            'VP': 'Vice President',
            'DIRECTOR': 'Division Manager',
            'AM': 'Area Manager',
            'FSM': 'Field Sales Manager'
        }
        level_title = level_titles.get(manager_level, manager_level)
        
        if test_email:
            mail.To = test_email
            mail.Subject = f"[TEST] Incentive Summary Report - {level_title} - {batch_month} {batch_year}"
        else:
            mail.To = recipient
            mail.Subject = f"Incentive Summary Report - {level_title} - {batch_month} {batch_year}"
            mail.CC = self.cc_email
        
        mail.SentOnBehalfOfName = self.sender_email
        
        mail.Body = ""  # Clear plain text body
        mail.HTMLBody = f"""
<html>
<body style="font-family: Calibri, Arial, sans-serif; font-size: 11pt; color: #333333;">

<p>{manager_name} ({manager_id}),</p>

<p>Attached is your team's incentive summary report for the <strong>{batch_month} {batch_year}</strong> payment period.</p>

<p>This report provides a detailed breakdown of incentive earnings for your team members, including:</p>
<ul style="margin-left: 20px;">
    <li>Summary tabs showing total payouts by team member</li>
    <li>Detail tabs with individual incentive line items</li>
</ul>

<p>Please review the report to understand your team's incentive performance. If you have questions about the calculations or need additional information, please contact the CA Analytics Team at <a href="mailto:CAAnalyticsAll@breakthrubev.com" style="color: #0563C1;">CAAnalyticsAll@breakthrubev.com</a></p>

<p style="margin-top: 20px;">-Analytics Team</p>

</body>
</html>
"""
        
        mail.Attachments.Add(str(attachment_path))
        
        return mail
    
    def send_mgmt_test_emails(self):
        """Send test management emails."""
        if not OUTLOOK_AVAILABLE:
            messagebox.showerror("Error", "Outlook not available")
            return
        
        test_email = self.mgmt_test_email_var.get().strip()
        if not test_email or '@' not in test_email:
            messagebox.showinfo("Test Email Required", "Please enter a valid test email address")
            return
        
        if not self.mgmt_email_mapping:
            messagebox.showinfo("No Mappings", "Run email assessment first")
            return
        
        try:
            outlook = win32.Dispatch('Outlook.Application')
            
            # Select up to 3 random managers
            sample_ids = random.sample(
                list(self.mgmt_email_mapping.keys()),
                min(3, len(self.mgmt_email_mapping))
            )
            
            batch_month = self.batch_var.get()
            batch_year = self.year_var.get()
            
            sent = 0
            for mgr_id in sample_ids:
                info = self.mgmt_email_mapping[mgr_id]
                
                mail = self.create_mgmt_email(
                    outlook, info['email'], mgr_id, info['name'], info['level'],
                    batch_month, batch_year, info['filepath'], test_email=test_email
                )
                mail.Send()
                sent += 1
            
            messagebox.showinfo("Test Sent", f"Sent {sent} test emails to {test_email}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to send test emails:\n{str(e)}")
    
    def send_mgmt_final_emails(self):
        """Send final management emails."""
        if not OUTLOOK_AVAILABLE:
            messagebox.showerror("Error", "Outlook not available")
            return
        
        if not self.mgmt_email_mapping:
            messagebox.showinfo("No Mappings", "Run email assessment first")
            return
        
        count = len(self.mgmt_email_mapping)
        if not messagebox.askyesno("Confirm Send", f"Send {count} management report emails?"):
            return
        
        try:
            outlook = win32.Dispatch('Outlook.Application')
            
            batch_month = self.batch_var.get()
            batch_year = self.year_var.get()
            
            self.stage3_progress['maximum'] = count
            self.stage3_progress['value'] = 0
            
            sent_emails = []
            failed_emails = []
            
            for i, (mgr_id, info) in enumerate(self.mgmt_email_mapping.items()):
                try:
                    mail = self.create_mgmt_email(
                        outlook, info['email'], mgr_id, info['name'], info['level'],
                        batch_month, batch_year, info['filepath']
                    )
                    mail.Send()
                    
                    sent_emails.append({
                        'ManagerID': mgr_id,
                        'Name': info['name'],
                        'Level': info['level'],
                        'Email': info['email']
                    })
                    
                except Exception as e:
                    failed_emails.append({
                        'ManagerID': mgr_id,
                        'Name': info['name'],
                        'Level': info['level'],
                        'Status': str(e)
                    })
                
                self.stage3_progress['value'] = i + 1
                self.stage3_progress_label.config(text=f"Sending {i+1} of {count}...")
                self.root.update()
            
            self.mgmt_send_status.config(
                text=f"✓ Sent {len(sent_emails)}, Failed {len(failed_emails)}",
                style="Success.TLabel" if not failed_emails else "Warning.TLabel"
            )
            self.stage3_progress_label.config(text=f"Complete: {len(sent_emails)} sent, {len(failed_emails)} failed")
            
            # Send recap
            self.send_mgmt_recap_email(outlook, sent_emails, failed_emails, batch_month, batch_year)
            
            messagebox.showinfo("Complete", f"Sent {len(sent_emails)} management emails")
            
        except Exception as e:
            messagebox.showerror("Error", f"Failed to send emails:\n{str(e)}")
    
    def send_mgmt_recap_email(self, outlook, sent_emails, failed_emails, batch_month, batch_year):
        """Send recap email for management report distribution."""
        try:
            mail = outlook.CreateItem(0)
            mail.To = "; ".join(self.recap_recipients)
            mail.Subject = f"Management Incentive Report Distribution Recap - {batch_month} {batch_year}"
            mail.SentOnBehalfOfName = self.sender_email
            
            total_sent = len(sent_emails)
            total_failed = len(failed_emails)
            success_rate = (total_sent / (total_sent + total_failed) * 100) if (total_sent + total_failed) > 0 else 0
            
            body = f"""
MANAGEMENT INCENTIVE REPORT DISTRIBUTION RECAP
{'='*50}

Batch: {batch_month} {batch_year}
Date Sent: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

SUMMARY
{'-'*30}
Total Emails Sent: {total_sent}
Total Failed: {total_failed}
Success Rate: {success_rate:.1f}%


DETAILED SEND LOG
{'-'*30}
"""
            
            # Group by level
            for level in ['VP', 'DIRECTOR', 'AM', 'FSM']:
                level_sent = [e for e in sent_emails if e['Level'] == level]
                if level_sent:
                    body += f"\n{level}s:\n"
                    for item in level_sent:
                        body += f"  ✓ {item['ManagerID']} | {item['Name']} | {item['Email']}\n"
            
            if failed_emails:
                body += f"""

FAILED SENDS
{'-'*30}
"""
                for item in failed_emails:
                    body += f"✗ {item['ManagerID']} | {item['Name']} | {item['Status']}\n"
            
            body += f"""

{'='*50}
Automated recap from Incentive Batch Manager
"""
            
            mail.Body = body
            mail.Send()
            
        except Exception as e:
            messagebox.showwarning("Recap Warning", f"Emails sent but recap failed:\n{str(e)}")


def main():
    root = tk.Tk()
    
    window_width = 750
    window_height = 1100
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width - window_width) // 2
    y = (screen_height - window_height) // 2
    root.geometry(f"{window_width}x{window_height}+{x}+{y}")
    
    app = IncentiveBatchManager(root)
    root.mainloop()


if __name__ == "__main__":
    main()